"""
Alberta Business & Data Analyst Job Scraper
Uses SerpAPI to scrape Google Jobs listings and saves results to Excel.

Requirements:
    pip install requests openpyxl

Usage:
    1. Set your SerpAPI key below (SERPAPI_KEY)
    2. Run: python alberta_job_scraper.py
    3. Output: alberta_jobs.xlsx
"""

import re
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import Counter, defaultdict
import time

# ─────────────────────────────────────────────
# CONFIGURATION — update your API key here
# ─────────────────────────────────────────────
SERPAPI_KEY = "YOUR_SERPAPI_KEY_HERE"

SEARCH_QUERIES = [
    "Business Analyst Alberta",
    "Senior Business Analyst Alberta",
    "Data Analyst Alberta",
    "Senior Data Analyst Alberta",
    "Business Intelligence Analyst Alberta",
    "Senior Business Intelligence Analyst Alberta",
    "BI Analyst Alberta",
    "Senior BI Analyst Alberta",
    "Financial Analyst Alberta",
    "Senior Financial Analyst Alberta",
    "Operations Analyst Alberta",
    "Senior Operations Analyst Alberta",
    "Reporting Analyst Alberta",
    "Senior Reporting Analyst Alberta",
    "Customer Insights Analyst Alberta",
    "Senior Customer Insights Analyst Alberta",
]

OUTPUT_FILE = "alberta_jobs.xlsx"

# Pagination settings
# Google Jobs returns exactly 10 results per page.
# MAX_PAGES_PER_QUERY controls how many pages to fetch per query.
# Each page costs 1 SerpAPI credit.
#   5 pages × 16 queries = 80 credits per full run
#  10 pages × 16 queries = 160 credits per full run
MAX_PAGES_PER_QUERY   = 5    # increase for more results (each page = 10 jobs)
DELAY_BETWEEN_REQUESTS = 1.0  # seconds between every API call


# ─────────────────────────────────────────────
# TITLE CLASSIFICATION
# ─────────────────────────────────────────────

TITLE_PATTERNS = [
    # ── Specific / compound titles first (order matters — most specific → least) ──
    (r"customer insight",                              "Customer Insights Analyst"),
    (r"business intelligence|bi analyst|b\.i\.",       "BI Analyst"),
    (r"fp&a|fp &a|financial planning.{0,10}analysis",  "FP&A Analyst"),
    (r"it business analyst|technical business analyst|it ba", "IT Business Analyst"),
    (r"supply chain analyst",                          "Supply Chain Analyst"),
    (r"commercial analyst",                            "Commercial Analyst"),
    (r"land analyst",                                  "Land Analyst"),
    (r"risk analyst",                                  "Risk Analyst"),
    (r"pricing analyst",                               "Pricing Analyst"),
    (r"data scientist",                                "Data Scientist"),
    (r"data analyst",                                  "Data Analyst"),
    (r"financial analyst",                             "Financial Analyst"),
    (r"operations analyst",                            "Operations Analyst"),
    (r"reporting analyst",                             "Reporting Analyst"),
    (r"systems analyst",                               "Systems Analyst"),
    # ── Business Analyst last so IT BA / FP&A are caught first ───────────────────
    (r"business analyst",                              "Business Analyst"),
]

SENIORITY_PATTERNS = [
    (r"\b(lead|principal|staff)\b",           "Lead"),
    (r"\b(senior|sr\.?|sr )\b",              "Senior"),
    (r"\b(junior|jr\.?|jr |entry.?level)\b", "Junior"),
    (r"\b(manager|head of|director)\b",       "Manager"),
]

def classify_title(raw_title: str) -> tuple[str, str]:
    lower = raw_title.lower()
    seniority = "Mid"
    for pattern, label in SENIORITY_PATTERNS:
        if re.search(pattern, lower):
            seniority = label
            break
    title_short = "Other Analyst"
    for pattern, label in TITLE_PATTERNS:
        if re.search(pattern, lower):
            title_short = label
            break
    # Prefix with seniority when not Mid
    # e.g. "Senior Business Analyst", "Lead BI Analyst", "Junior Data Analyst"
    if seniority != "Mid":
        title_short = f"{seniority} {title_short}"
    return title_short, seniority


# ─────────────────────────────────────────────
# WORK MODE  (Remote / Hybrid / On-Site)
# Checks both the SerpAPI work_from_home flag AND description text
# ─────────────────────────────────────────────

def detect_work_mode(description: str, wfh_flag: bool) -> str:
    """
    Returns 'Remote', 'Hybrid', or 'On-Site'.
    Hybrid takes priority over pure Remote when both signals exist.
    """
    text = description.lower()

    is_hybrid = bool(re.search(
        r'\bhybrid\b|hybrid.work|hybrid.model|hybrid.schedule|'
        r'combination of (remote|in.office)|partial.remote|'
        r'(2|3|two|three).days?.*(office|on.?site)|'
        r'(office|on.?site).*(2|3|two|three).days?',
        text
    ))
    is_remote = bool(re.search(
        r'\bremote\b|work from home|work.from.home|'
        r'fully remote|100%.remote|telecommut|telework|'
        r'anywhere in (canada|alberta)',
        text
    )) or wfh_flag

    if is_hybrid:
        return "Hybrid"
    if is_remote:
        return "Remote"
    return "On-Site"


# ─────────────────────────────────────────────
# EDUCATION / CERTIFICATION EXTRACTION
# ─────────────────────────────────────────────

# Ordered from highest to lowest — first match wins for degree level
DEGREE_PATTERNS = [
    (r"\b(ph\.?d\.?|doctorate|doctoral)\b",                              "PhD"),
    (r"\b(master'?s?|m\.?sc\.?|m\.?b\.?a\.?|m\.?eng\.?|graduate degree)\b", "Master's"),
    (r"\b(bachelor'?s?|b\.?sc\.?|b\.?a\.?|b\.?eng\.?|b\.?comm\.?|undergraduate degree|university degree|4.year degree)\b", "Bachelor's"),
    (r"\b(diploma|college diploma|post.secondary)\b",                    "Diploma"),
    (r"\b(certificate|certification|cert\.)\b",                          "Certificate"),
    (r"\b(high school|secondary school|ged)\b",                          "High School"),
]

# Specific certifications to surface (searched separately, appended)
CERT_PATTERNS = [
    r"\bPMP\b", r"\bCBP\b", r"\bCBP\b", r"\bCPB\b", r"\bCBA\b",
    r"\bCPA\b", r"\bCFA\b", r"\bCMA\b", r"\bCFP\b",
    r"\bIIBA\b", r"\bECBA\b", r"\bCCBA\b", r"\bCBAP\b",
    r"\bCSPO\b", r"\bCSM\b", r"\bSAFe\b",
    r"\bAWS Certified\b", r"\bAzure Certified\b", r"\bGoogle Cloud\b",
    r"\bTableau Desktop Specialist\b", r"\bTableau Certified\b",
    r"\bPower BI Certified\b", r"\bMicrosoft Certified\b",
    r"\bSix Sigma\b", r"\bLean\b",
    r"\bData Science Certificate\b", r"\bData Analytics Certificate\b",
]
_CERT_RX = [re.compile(p, re.IGNORECASE) for p in CERT_PATTERNS]

# Common study fields to capture
FIELD_PATTERNS = [
    (r"computer science|computing science",           "Computer Science"),
    (r"information (technology|systems|management)",  "Information Technology"),
    (r"data science|data analytics",                  "Data Science/Analytics"),
    (r"business administration|business management",  "Business Administration"),
    (r"business (analytics|intelligence)",            "Business Analytics/BI"),
    (r"finance|accounting|economics",                 "Finance/Accounting"),
    (r"mathematics|statistics|quantitative",          "Math/Statistics"),
    (r"engineering",                                  "Engineering"),
    (r"operations research|management science",       "Operations Research"),
    (r"related field|equivalent",                     "Related Field"),
]

def extract_education(description: str) -> str:
    """
    Returns a compact string like: 'Bachelor's (Computer Science) | PMP, CBAP'
    Components that are not found are simply omitted.
    """
    text = description.lower()
    parts = []

    # Degree level
    degree = ""
    for pattern, label in DEGREE_PATTERNS:
        if re.search(pattern, text):
            degree = label
            break

    # Field of study
    field = ""
    for pattern, label in FIELD_PATTERNS:
        if re.search(pattern, text):
            field = label
            break

    if degree and field:
        parts.append(f"{degree} ({field})")
    elif degree:
        parts.append(degree)
    elif field:
        parts.append(f"Degree in {field}")

    # Named certifications
    certs = [rx.pattern.strip(r'\b').replace(r'\b', '') for rx in _CERT_RX if rx.search(description)]
    # Clean up regex artifacts for display
    certs_clean = []
    for p, rx in zip(CERT_PATTERNS, _CERT_RX):
        if rx.search(description):
            # Use the readable part of the pattern
            readable = re.sub(r'\\b|\\', '', p).strip()
            certs_clean.append(readable)

    if certs_clean:
        parts.append(", ".join(certs_clean))

    return " | ".join(parts) if parts else ""


def classify_education(description: str) -> tuple[str, str]:
    """
    Returns (edu_level, edu_field) — clean, filterable short labels,
    analogous to (title_short, seniority) for job titles.

    edu_level : 'PhD' | "Master's" | "Bachelor's" | 'Diploma' |
                'Certificate' | 'High School' | ''
    edu_field : 'Computer Science' | 'Business Administration' | ... | ''
    """
    text = description.lower()

    edu_level = ""
    for pattern, label in DEGREE_PATTERNS:
        if re.search(pattern, text):
            edu_level = label
            break

    edu_field = ""
    for pattern, label in FIELD_PATTERNS:
        if re.search(pattern, text):
            edu_field = label
            break

    return edu_level, edu_field



# ─────────────────────────────────────────────
# SALARY PARSING
# Handles formats like:
#   "$80,000–$100,000 a year"
#   "$80K–$100K a year"
#   "$80k/year"
#   "From $90K a year"
#   "$40–$55 an hour"
#   "$40.50/hr"
#   "$75,000 a year"
# ─────────────────────────────────────────────

_HOURLY_RX = re.compile(r'\b(hour|hr|hourly|\/h)\b', re.IGNORECASE)
_YEARLY_RX = re.compile(r'\b(year|yr|annual|annually|salary)\b', re.IGNORECASE)

# Matches: $80K, $80.5K, $80,000, $80,000.50, $80  (with optional K/k suffix)
_SALARY_NUMBER_RX = re.compile(r'\$[\d,]+(?:\.\d+)?[Kk]?')

def _expand_k(token: str) -> float:
    """Convert a salary token like '$80K' or '$80,000' to a float."""
    clean = token.replace("$", "").replace(",", "")
    if clean.upper().endswith("K"):
        return float(clean[:-1]) * 1000
    return float(clean)

def parse_salary(raw_salary: str) -> dict:
    """
    Returns dict with keys:
        salary_type  : 'Hourly' | 'Yearly' | ''
        salary_min   : float or ''
        salary_max   : float or ''
        salary_avg   : float or ''
    """
    result = {"salary_type": "", "salary_min": "", "salary_max": "", "salary_avg": ""}
    if not raw_salary:
        return result

    # Determine pay period
    if _HOURLY_RX.search(raw_salary):
        result["salary_type"] = "Hourly"
    elif _YEARLY_RX.search(raw_salary):
        result["salary_type"] = "Yearly"

    # Extract and expand all dollar amounts (handles $80K → 80000)
    amounts = []
    for m in _SALARY_NUMBER_RX.finditer(raw_salary):
        try:
            amounts.append(_expand_k(m.group()))
        except ValueError:
            pass

    if len(amounts) == 0:
        return result
    elif len(amounts) == 1:
        result["salary_min"] = amounts[0]
        result["salary_max"] = amounts[0]
        result["salary_avg"] = amounts[0]
    else:
        result["salary_min"] = min(amounts)
        result["salary_max"] = max(amounts)
        result["salary_avg"] = round((result["salary_min"] + result["salary_max"]) / 2, 2)

    return result


# ─────────────────────────────────────────────
# CITY EXTRACTION
# Parses the raw Google Jobs location string into "City, AB"
# Handles formats like:
#   "Calgary, AB"              → "Calgary, AB"
#   "Calgary, AB +1 other"    → "Calgary, AB"
#   "Edmonton, Alberta"        → "Edmonton, AB"
#   "Alberta, Canada"          → ""   (no specific city)
#   "Alberta"                  → ""
# ─────────────────────────────────────────────

# Province-level placeholders that are NOT a city — set to null
_PROVINCE_ONLY = {
    "alberta", "ab", "alberta, canada", "alberta, ab", "canada",
}

# Normalize known province suffixes to ", AB"
_PROVINCE_SUFFIX_RX = re.compile(
    r',?\s*(alberta|ab|alberta,?\s*canada)\b.*$',
    re.IGNORECASE
)

def extract_city(raw_location: str) -> str:
    """
    Returns 'City, AB' from a raw Google Jobs location string,
    or '' if no specific city can be determined.
    """
    if not raw_location:
        return ""

    # Strip trailing noise like "+1 other", "+2 locations", "(Remote)", etc.
    cleaned = re.sub(
        r'\s*[\+\(].*$',          # everything from + or ( onwards
        '',
        raw_location,
        flags=re.IGNORECASE
    ).strip().rstrip(',').strip()

    # If what remains is just a province/country placeholder → null
    if cleaned.lower() in _PROVINCE_ONLY:
        return ""

    # If it starts with "Alberta" and has no meaningful city prefix → null
    if re.match(r'^alberta\b', cleaned, re.IGNORECASE):
        return ""

    # Normalise the province suffix: "Edmonton, Alberta" → "Edmonton, AB"
    city_part = _PROVINCE_SUFFIX_RX.sub('', cleaned).strip().rstrip(',').strip()

    # If city_part is empty or still looks like a province only → null
    if not city_part or city_part.lower() in _PROVINCE_ONLY:
        return ""

    return f"{city_part}, AB"


# ─────────────────────────────────────────────
# INDUSTRY DETECTION
# ─────────────────────────────────────────────

INDUSTRY_RULES = [
    (r"oil|gas|pipeline|petroleum|energy|suncor|cenovus|cnrl|enbridge|tc energy|pembina|atco|encana|ovintiv|repsol|husky",
     "Oil & Gas / Energy"),
    (r"bank|financ|insurance|credit union|invest|wealth|mortgage|atb|rbc|td |bmo|scotiabank|cibc|manulife|sunlife|intact",
     "Finance & Banking"),
    (r"health|hospital|clinic|pharma|medical|dental|alberta health|ahs|covenant|dynalife|telus health",
     "Healthcare"),
    (r"tech|software|saas|cloud|cyber|data platform|it |information technology|telecom|telus|shaw|rogers|ericsson",
     "Technology & Telecom"),
    (r"retail|grocery|wholesale|e.commerce|consumer|sobeys|loblaws|canadian tire|walmart|costco|amazon",
     "Retail & Consumer"),
    (r"government|municipal|province|public sector|crown|city of|town of|county|alberta government|ministry",
     "Government & Public Sector"),
    (r"consult|advisory|deloitte|kpmg|pwc|ernst|mckinsey|accenture|capgemini|mca",
     "Consulting"),
    (r"transport|logistics|supply chain|freight|shipping|fleet|trucking|rail|aviation|air canada|westjet",
     "Transportation & Logistics"),
    (r"construction|real estate|infrastructure|engineer|aecom|stantec|wsp|jacobs|pcl|graham",
     "Construction & Engineering"),
    (r"education|university|college|school|academic|ucalgary|ualberta|nait|sait|bow valley",
     "Education"),
    (r"utility|power|electric|enmax|epcor|fortis|nova|gas distribution|water",
     "Utilities"),
    (r"manufactur|industrial|plant|production|chemical|agri|food processing",
     "Manufacturing & Industrial"),
    (r"media|marketing|advertising|communications|pr |public relations|digital agency",
     "Media & Marketing"),
    (r"non.profit|ngo|charity|social service|community|foundation",
     "Non-Profit & Social Services"),
]

def detect_industry(company: str, description: str) -> str:
    text = (company + " " + description).lower()
    for pattern, label in INDUSTRY_RULES:
        if re.search(pattern, text):
            return label
    return "Other"


# ─────────────────────────────────────────────
# SKILLS EXTRACTION
# ─────────────────────────────────────────────

SKILLS_LIST = [
    "SQL", "Python", "R", "Excel", "Power BI", "Tableau", "Looker", "Qlik",
    "SSRS", "SSIS", "SSAS", "Azure", "AWS", "GCP", "Databricks", "Snowflake",
    "dbt", "Spark", "Hadoop", "Hive", "Alteryx", "SAS", "SPSS", "MATLAB",
    "DAX", "MDX", "Power Query", "Power Automate", "Power Apps",
    "Google Analytics", "Adobe Analytics", "Salesforce", "SAP", "Oracle",
    "Workday", "Dynamics 365", "ServiceNow", "Jira", "Confluence",
    "SharePoint", "Visio", "Lucidchart",
    "Financial Modeling", "Budgeting", "Forecasting", "Variance Analysis",
    "P&L", "Cash Flow", "GAAP", "IFRS", "Hyperion", "Anaplan", "TM1",
    "Business Objects", "MicroStrategy",
    "Agile", "Scrum", "Kanban", "Waterfall", "SDLC", "Six Sigma", "Lean",
    "BPMN", "UML", "ETL", "Data Warehousing", "Data Modeling",
    "Machine Learning", "Statistics", "A/B Testing", "Data Governance",
    "KPIs", "Dashboards", "Reporting", "Data Visualization",
    "Requirements Gathering", "User Stories", "Process Mapping",
    "Stakeholder Management", "Change Management",
    "Communication", "Presentation", "Problem Solving", "Critical Thinking",
    "Collaboration", "Leadership",
]

_SKILL_PATTERNS = [
    (s, re.compile(r'\b' + re.escape(s) + r'\b', re.IGNORECASE))
    for s in SKILLS_LIST
]

def extract_skills(description: str) -> str:
    return ", ".join(skill for skill, pat in _SKILL_PATTERNS if pat.search(description))


# ─────────────────────────────────────────────
# SCRAPER
# ─────────────────────────────────────────────

def _fetch_page(query: str, api_key: str, next_page_token: str = None) -> tuple[list, str | None]:
    """
    Fetch one page of Google Jobs results via SerpAPI.
    Returns (jobs_list, next_page_token_or_None).

    Google Jobs pagination uses next_page_token — the old `start` offset
    parameter was discontinued by Google and causes a 400 error.
    """
    url = "https://serpapi.com/search"
    params = {
        "engine":   "google_jobs",
        "q":        query,
        "location": "Alberta, Canada",
        "hl":       "en",
        "gl":       "ca",
        "api_key":  api_key,
    }
    if next_page_token:
        params["next_page_token"] = next_page_token

    try:
        response = requests.get(url, params=params, timeout=20)
        response.raise_for_status()
        data = response.json()

        if "error" in data:
            print(f"    [API ERROR] {data['error']}")
            return [], None

        jobs       = data.get("jobs_results", [])
        pagination = data.get("serpapi_pagination", {})
        token      = pagination.get("next_page_token")   # None on last page
        return jobs, token

    except requests.exceptions.RequestException as e:
        print(f"    [REQUEST ERROR] {e}")
        return [], None


def _parse_job(job: dict, query: str) -> dict:
    """Enrich a raw SerpAPI job dict into our output schema."""
    apply_link = ""
    apply_options = job.get("apply_options", [])
    if apply_options:
        apply_link = apply_options[0].get("link", "")

    extensions     = job.get("detected_extensions", {})
    raw_salary     = extensions.get("salary", "")
    job_type       = extensions.get("schedule_type", "")
    posted_date    = extensions.get("posted_at", "")
    work_from_home = extensions.get("work_from_home", False)

    raw_title   = job.get("title", "")
    company     = job.get("company_name", "")
    description = job.get("description", "")

    title_short, seniority   = classify_title(raw_title)
    industry                  = detect_industry(company, description)
    skills                    = extract_skills(description)
    work_mode                 = detect_work_mode(description, work_from_home)
    education                 = extract_education(description)
    edu_level, edu_field      = classify_education(description)
    sal                       = parse_salary(raw_salary)
    raw_loc                   = job.get("location", "")
    city                      = extract_city(raw_loc)

    return {
        "Job Title":      raw_title,
        "Title (Short)":  title_short,
        "Seniority":      seniority,
        "Company":        company,
        "Industry":       industry,
        "Location":       raw_loc,
        "City":           city,
        "Work Mode":      work_mode,
        "Job Type":       job_type,
        "Salary (Raw)":   raw_salary,
        "Salary Type":    sal["salary_type"],
        "Salary Min":     sal["salary_min"],
        "Salary Max":     sal["salary_max"],
        "Salary Avg":     sal["salary_avg"],
        "Education":      education,
        "Edu Level":      edu_level,
        "Edu Field":      edu_field,
        "Skills":         skills,
        "Has Work Mode":  bool(work_mode and work_mode != "On-Site"),
        "Has Job Type":   bool(job_type),
        "Has Salary":     bool(raw_salary),
        "Has Education":  bool(education),
        "Has Skills":     bool(skills),
        "Date Posted":    posted_date,
        "Description":    description[:700].replace("\n", " "),
        "Apply Link":     apply_link,
        "Search Query":   query,
        "Scraped At":     datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


def scrape_all_jobs(queries: list[str], api_key: str) -> list[dict]:
    """
    Iterates over every query and paginates using next_page_token
    (Google Jobs\'s only supported pagination method).
    Stops when:
      - no next_page_token is returned  (end of results), OR
      - MAX_PAGES_PER_QUERY pages have been fetched for this query.
    Deduplicates across all queries by (title, company).
    """
    all_jobs       = []
    seen           = set()   # (title_lower, company_lower)
    total_api_calls = 0

    for q_idx, query in enumerate(queries, 1):
        query_new  = 0
        token      = None    # first page has no token
        print(f"[{q_idx}/{len(queries)}] {query}")

        for page in range(1, MAX_PAGES_PER_QUERY + 1):
            print(f"  page {page}/{MAX_PAGES_PER_QUERY} ...", end=" ", flush=True)

            raw_jobs, token = _fetch_page(query, api_key, next_page_token=token)
            total_api_calls += 1
            print(f"{len(raw_jobs)} results")

            for raw in raw_jobs:
                parsed = _parse_job(raw, query)
                key = (parsed["Job Title"].lower().strip(),
                       parsed["Company"].lower().strip())
                if key not in seen:
                    seen.add(key)
                    all_jobs.append(parsed)
                    query_new += 1

            # No token means Google has no more pages for this query
            if not token:
                print(f"  → no more pages")
                break

            time.sleep(DELAY_BETWEEN_REQUESTS)

        print(f"  → {query_new} new unique jobs added  (total so far: {len(all_jobs)})")

        if q_idx < len(queries):
            time.sleep(DELAY_BETWEEN_REQUESTS)

    print(f"\nTotal API calls used : {total_api_calls}")
    print(f"Total unique jobs    : {len(all_jobs)}")
    return all_jobs


# ─────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────

COLUMNS = [
    # ── Identity ──────────────────────
    "Job Title",
    "Title (Short)",
    "Seniority",
    "Company",
    "Industry",
    "Location",
    "City",
    # ── Work conditions ───────────────
    "Work Mode",
    "Job Type",
    # ── Salary (parsed) ───────────────
    "Salary (Raw)",
    "Salary Type",
    "Salary Min",
    "Salary Max",
    "Salary Avg",
    # ── Requirements ──────────────────
    "Education",
    "Edu Level",
    "Edu Field",
    "Skills",
    # ── Data completeness flags ───────
    "Has Work Mode",
    "Has Job Type",
    "Has Salary",
    "Has Education",
    "Has Skills",
    # ── Content ───────────────────────
    "Date Posted",
    "Description",
    "Apply Link",
    "Search Query",
    "Scraped At",
]

# Colour palette
HEADER_COLOR  = "1F4E79"   # dark navy
ACCENT_COLOR  = "2E75B6"   # medium blue  (enriched columns)
GREEN_COLOR   = "375623"   # dark green   (salary columns)
ALT_ROW       = "D6E4F0"   # pale blue alternating rows
LINK_COLOR    = "0563C1"
GREEN_ALT     = "E2EFDA"   # pale green alternating rows for salary

ENRICHED_COLS = {"Title (Short)", "Seniority", "Industry", "City", "Work Mode", "Education", "Edu Level", "Edu Field", "Skills"}
SALARY_COLS   = {"Salary Type", "Salary Min", "Salary Max", "Salary Avg"}
FLAG_COLS     = {"Has Work Mode", "Has Job Type", "Has Salary", "Has Education", "Has Skills"}
ORANGE_COLOR  = "833C00"   # dark orange  (flag columns)
ORANGE_ALT    = "FCE4D6"   # pale orange  (flag alternating rows)

COL_WIDTHS = {
    "Job Title":     38,  "Title (Short)": 22,  "Seniority":     12,
    "Company":       26,  "Industry":      24,  "Location":      22,  "City":          18,
    "Work Mode":     12,  "Job Type":      14,
    "Salary (Raw)":  24,  "Salary Type":   12,
    "Salary Min":    14,  "Salary Max":    14,  "Salary Avg":    14,
    "Education":     36,  "Edu Level":     16,  "Edu Field":     26,
    "Skills":        55,
    "Has Work Mode": 14,  "Has Job Type":  12,  "Has Salary":    12,
    "Has Education": 14,  "Has Skills":    12,
    "Date Posted":   14,  "Description":   60,
    "Apply Link":    12,  "Search Query":  32,  "Scraped At":    18,
}


def _header_color_for(col_name: str) -> str:
    if col_name in SALARY_COLS:
        return GREEN_COLOR
    if col_name in FLAG_COLS:
        return ORANGE_COLOR
    if col_name in ENRICHED_COLS:
        return ACCENT_COLOR
    return HEADER_COLOR


def style_header(cell, col_name=""):
    color = _header_color_for(col_name)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFFFFF")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_data_cell(cell, row_is_even, col_name="", is_link=False):
    is_salary = col_name in SALARY_COLS
    is_flag   = col_name in FLAG_COLS
    is_tag    = col_name in ENRICHED_COLS

    if is_salary:
        bg = GREEN_ALT if row_is_even else "FFFFFF"
    elif is_flag:
        bg = ORANGE_ALT if row_is_even else "FFFFFF"
    else:
        bg = ALT_ROW if row_is_even else "FFFFFF"

    cell.fill = PatternFill("solid", fgColor=bg)

    if is_link:
        color, underline = LINK_COLOR, "single"
    elif is_salary:
        color, underline = "375623", None
    elif is_flag:
        # True = dark green text, False = red text
        color = "375623" if cell.value is True else "C00000"
        underline = None
    elif is_tag:
        color, underline = ACCENT_COLOR, None
    else:
        color, underline = "000000", None

    if not is_flag:  # font set below for flags after value check
        cell.font = Font(name="Arial", size=10, color=color, underline=underline)
    else:
        cell.font = Font(name="Arial", size=10, bold=True,
                         color="375623" if cell.value is True else "C00000")
    cell.alignment = Alignment(
        vertical="top",
        wrap_text=False,
        horizontal="center" if is_flag else ("right" if is_salary and col_name != "Salary Type" else "left")
    )
    thin = Side(style="thin", color="D9D9D9")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Number formatting for salary values
    if col_name in ("Salary Min", "Salary Max", "Salary Avg") and isinstance(cell.value, (int, float)):
        cell.number_format = '#,##0.00'


def _write_jobs_to_sheet(ws, jobs, cols):
    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        style_header(cell, col_name)
    ws.row_dimensions[1].height = 32

    for row_idx, job in enumerate(jobs, 2):
        is_even = (row_idx % 2 == 0)
        for col_idx, col_name in enumerate(cols, 1):
            value = job.get(col_name, "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            is_link = (col_name == "Apply Link" and str(value).startswith("http"))
            style_data_cell(cell, is_even, col_name=col_name, is_link=is_link)
            if is_link:
                cell.hyperlink = value
                cell.value = "Apply →"

    for col_idx, col_name in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 20)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def _hdr_row(ws, row, *vals, color=None):
    c = color or HEADER_COLOR
    for col, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cell.fill  = PatternFill("solid", fgColor=c)
        cell.alignment = Alignment(horizontal="center")


def _dat_row(ws, row, *vals):
    for col, v in enumerate(vals, 1):
        ws.cell(row=row, column=col, value=v).font = Font(name="Arial", size=10)


def save_to_excel(jobs: list[dict], filename: str):
    wb = openpyxl.Workbook()

    # ── Summary sheet ──────────────────────────────────────────
    ws_s = wb.active
    ws_s.title = "Summary"

    ws_s.merge_cells("A1:D1")
    t = ws_s["A1"]
    t.value = "Alberta Analyst Job Scraper — Results"
    t.font  = Font(name="Arial", bold=True, size=14, color=HEADER_COLOR)
    t.alignment = Alignment(horizontal="center")
    ws_s.row_dimensions[1].height = 32

    cur = 3

    # Overall stats
    _hdr_row(ws_s, cur, "Metric", "Value")
    cur += 1
    for label, val in [
        ("Total Unique Jobs", len(jobs)),
        ("Queries Run",       len(SEARCH_QUERIES)),
        ("Date Scraped",      datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]:
        _dat_row(ws_s, cur, label, val); cur += 1

    cur += 1

    # By title short
    title_counts = Counter(j["Title (Short)"] for j in jobs)
    _hdr_row(ws_s, cur, "Title (Short)", "Count"); cur += 1
    for v, c2 in sorted(title_counts.items(), key=lambda x: -x[1]):
        _dat_row(ws_s, cur, v, c2); cur += 1
    cur += 1

    # By seniority
    sen_counts = Counter(j["Seniority"] for j in jobs)
    _hdr_row(ws_s, cur, "Seniority", "Count"); cur += 1
    for v, c2 in sorted(sen_counts.items(), key=lambda x: -x[1]):
        _dat_row(ws_s, cur, v, c2); cur += 1
    cur += 1

    # By industry
    ind_counts = Counter(j["Industry"] for j in jobs)
    _hdr_row(ws_s, cur, "Industry", "Count"); cur += 1
    for v, c2 in sorted(ind_counts.items(), key=lambda x: -x[1]):
        _dat_row(ws_s, cur, v, c2); cur += 1
    cur += 1

    # By work mode
    wm_counts = Counter(j["Work Mode"] for j in jobs)
    _hdr_row(ws_s, cur, "Work Mode", "Count", color=ACCENT_COLOR); cur += 1
    for v, c2 in sorted(wm_counts.items(), key=lambda x: -x[1]):
        _dat_row(ws_s, cur, v, c2); cur += 1
    cur += 1

    # Salary summary (yearly jobs only)
    yearly = [j for j in jobs if j.get("Salary Type") == "Yearly" and j.get("Salary Avg")]
    hourly = [j for j in jobs if j.get("Salary Type") == "Hourly" and j.get("Salary Avg")]
    _hdr_row(ws_s, cur, "Salary Insight", "Value", color=GREEN_COLOR); cur += 1
    if yearly:
        avgs = [j["Salary Avg"] for j in yearly]
        _dat_row(ws_s, cur, "Avg Yearly Salary (of posted)", f"${sum(avgs)/len(avgs):,.0f}"); cur += 1
        _dat_row(ws_s, cur, "Yearly Salary Range", f"${min(j['Salary Min'] for j in yearly):,.0f} – ${max(j['Salary Max'] for j in yearly):,.0f}"); cur += 1
    if hourly:
        avgs = [j["Salary Avg"] for j in hourly]
        _dat_row(ws_s, cur, "Avg Hourly Rate (of posted)", f"${sum(avgs)/len(avgs):,.2f}/hr"); cur += 1
        _dat_row(ws_s, cur, "Hourly Rate Range", f"${min(j['Salary Min'] for j in hourly):,.2f} – ${max(j['Salary Max'] for j in hourly):,.2f}/hr"); cur += 1
    _dat_row(ws_s, cur, "Jobs with Salary Posted", f"{len(yearly)+len(hourly)} / {len(jobs)}"); cur += 1
    cur += 1

    # By query
    q_counts = Counter(j["Search Query"] for j in jobs)
    _hdr_row(ws_s, cur, "Search Query", "Jobs Found"); cur += 1
    for v, c2 in sorted(q_counts.items(), key=lambda x: -x[1]):
        _dat_row(ws_s, cur, v, c2); cur += 1

    ws_s.column_dimensions["A"].width = 46
    ws_s.column_dimensions["B"].width = 26

    # ── All Jobs sheet ──────────────────────────────────────────
    ws_all = wb.create_sheet("All Jobs")
    _write_jobs_to_sheet(ws_all, jobs, COLUMNS)

    # ── Per-query sheets ────────────────────────────────────────
    per_query_cols = [c for c in COLUMNS if c != "Search Query"]
    jobs_by_query  = defaultdict(list)
    for job in jobs:
        jobs_by_query[job["Search Query"]].append(job)

    for query, query_jobs in jobs_by_query.items():
        sheet_name = (
            query[:31]
            .replace("/", "-").replace("\\", "-")
            .replace("?", "").replace("*", "")
            .replace("[", "").replace("]", "")
            .replace(":", "")
        )
        ws_q = wb.create_sheet(sheet_name)
        _write_jobs_to_sheet(ws_q, query_jobs, per_query_cols)

    wb.save(filename)
    print(f"✓ Results saved → {filename}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

if __name__ == "__main__":
    if SERPAPI_KEY == "YOUR_SERPAPI_KEY_HERE":
        print("ERROR: Please set your SerpAPI key in the SERPAPI_KEY variable at the top of this file.")
        exit(1)

    print("=" * 62)
    print("  Alberta Analyst Job Scraper  |  SerpAPI + Google Jobs")
    print("=" * 62)
    print(f"Queries : {len(SEARCH_QUERIES)}")
    print(f"Output  : {OUTPUT_FILE}\n")

    jobs = scrape_all_jobs(SEARCH_QUERIES, SERPAPI_KEY)

    if not jobs:
        print("No jobs found. Check your API key and internet connection.")
        exit(1)

    save_to_excel(jobs, OUTPUT_FILE)

    print("\nDone! Open alberta_jobs.xlsx to explore your results.")
    print("Sheets:")
    print("  • Summary   — stats by title, seniority, industry, work mode & salary")
    print("  • All Jobs  — full filterable table")
    print("  • Per query — one tab per search term")
